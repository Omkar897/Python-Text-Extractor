import cv2
import pytesseract
import re
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font

# Truncate text after the "Note" section
def truncate_after_note_section(text):
    note_section_pattern = re.compile(r'Note\s*[:-]\s*.*', re.IGNORECASE)
    large_gap_pattern = re.compile(r'\n\s*\n')

    match = note_section_pattern.search(text)
    if match:
        note_start = match.start()
        large_gap_match = large_gap_pattern.search(text[note_start:])
        if large_gap_match:
            truncated_text = text[:note_start + large_gap_match.start()]
        else:
            truncated_text = text[:note_start]
    else:
        truncated_text = text
    return truncated_text

# Extract patient information
def extract_patient_info(text):
    patient_info = {
        'Name': '',
        'SID': '',
        'Age': '',
        'Sex': ''
    }
    
    name_pattern = re.compile(r'^(?P<name>[A-Z\s]+)\s+Ref', re.MULTILINE)
    sid_pattern = re.compile(r'SID:\s*(?P<sid>\d+)', re.MULTILINE)
    age_sex_pattern = re.compile(r'Age:\s*(?P<age>[\d.]+)\s*Years\s*Sex:\s*(?P<sex>\w+)', re.IGNORECASE | re.MULTILINE)

    name_match = name_pattern.search(text)
    sid_match = sid_pattern.search(text)
    age_sex_match = age_sex_pattern.search(text)
    
    if name_match:
        patient_info['Name'] = name_match.group('name').strip()
    if sid_match:
        patient_info['SID'] = sid_match.group('sid').strip()
    if age_sex_match:
        patient_info['Age'] = age_sex_match.group('age').strip()
        patient_info['Sex'] = age_sex_match.group('sex').strip()
    
    return patient_info

# Extract and map sections
def extract_and_map_sections(text):
    markers = [
        "Clinical details", "Nature of specimen", 
        "Gross Examination", "Microscopy", "Diagnosis", 
        "HPE no.", "Note"
    ]
    
    key_value_pattern = re.compile(r'(\w[\w\s]*):')
    sections = {}
    current_key = None
    lines = text.splitlines()

    for line in lines:
        line = line.strip()
        match = key_value_pattern.match(line)
        if match:
            key = match.group(1).strip()
            # Clean and normalize the key
            key = key.replace(":", "").replace("  ", " ")
            if key in markers:
                current_key = key
                sections[current_key] = line.replace(match.group(0), '').strip()
            else:
                if current_key:
                    sections[current_key] += " " + line.strip()
        else:
            if current_key:
                sections[current_key] += " " + line.strip()

    if "Diagnosis" in sections:
        diagnosis = sections["Diagnosis"]
        hpe_match = re.search(r'HPE no\.?\s*[:;]?\s*(.+)', diagnosis, re.IGNORECASE)
        if hpe_match:
            sections["Diagnosis"] = diagnosis[:hpe_match.start()].strip()
            hpe_no = hpe_match.group(0).strip()
            sections["HPE no."] = hpe_no  
        elif "HPE no." in sections:
            sections["Diagnosis"] += f"\n\nHPE no.: {sections.pop('HPE no.')}"
    
    return sections

# Preprocess image to enhance OCR
def preprocess_image(image_path):
    image = cv2.imread(image_path)
    if image is None:
        raise ValueError(f"Could not open or find the image: {image_path}")
    
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    _, binary_image = cv2.threshold(gray, 150, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    return binary_image

def main(image_path):
    binary_image = preprocess_image(image_path)
    text = pytesseract.image_to_string(binary_image)
    
    patient_info = extract_patient_info(text)
    
    truncated_text = truncate_after_note_section(text)
    
    sections = extract_and_map_sections(truncated_text)
    
    reordered_text = "\n\n".join(f"{key}:\n{description}" for key, description in sections.items())
    
    print("Patient Information:\n")
    print(f"Name: {patient_info['Name']}\nSID: {patient_info['SID']}\nAge: {patient_info['Age']}\nSex: {patient_info['Sex']}\n")
    
    print("\nReordered Text:\n")
    print(reordered_text)

    data = [
        ('Name', patient_info['Name']),
        ('SID', patient_info['SID']),
        ('Age', patient_info['Age']),
        ('Sex', patient_info['Sex'])
    ] + list(sections.items())

    # Openpyxl Workbook setup
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Patient Report'

    max_columns = 6  

    # Write headers
    sheet['A1'] = 'Key'
    sheet['B1'] = 'Value'
    sheet['A1'].font = Font(bold=True)
    sheet['B1'].font = Font(bold=True)
    sheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=max_columns)
    sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
    sheet['B1'].alignment = Alignment(horizontal='center', vertical='center')

    # Write patient information and sections
    for i, (key, value) in enumerate(data, start=2):
        sheet[f'A{i}'] = key
        sheet[f'A{i}'].alignment = Alignment(horizontal='left', vertical='top')
        sheet.merge_cells(start_row=i, start_column=2, end_row=i, end_column=max_columns)
        sheet.cell(row=i, column=2, value=value).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        sheet.row_dimensions[i].height = max((len(value) // 50 + 1) * 15, 15)  
    
    # Adjust column widths
    for col in range(1, max_columns + 1):
        column_letter = get_column_letter(col)
        sheet.column_dimensions[column_letter].width = 20 if col == 1 else 50

    # Save the workbook
    excel_filename = image_path.replace('.jpg', '.xlsx')
    workbook.save(excel_filename)
    print(f"\nData exported to {excel_filename}")

if __name__ == "__main__":
    image_path = "img9.jpg"
    main(image_path)
