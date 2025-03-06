import streamlit as st
import cv2
import pytesseract
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
from PIL import Image
import numpy as np

def extract_text_from_image(image):
    """
    Extract text from an image using Tesseract OCR.
    """
    image_np = np.array(image)
    return pytesseract.image_to_string(image_np)

def extract_key_value_pairs(text):
    """
    Extract key-value pairs from the text based on known markers.
    """
    key_value_pairs = {}

    # Define regex patterns for known keys
    patterns = {
        "Patient Name": re.compile(r'Patient Name\s*:\s*(.*?)(?=\s*(Reference No|Clinical details|$))', re.S),
        "Reference No": re.compile(r'Reference No\s*:\s*(.*?)(?=\s*(Clinical details|$))', re.S),
        "Clinical details": re.compile(r'Clinical details\s*:\s*(.*?)(?=\s*(Nature of specimen|Gross Examination|Microscopy|Diagnosis|HPE no|Note|$))', re.S),
        "Nature of specimen": re.compile(r'Nature ofspecimen\s*:\s*(.*?)(?=\s*(Gross Examination|Microscopy|Diagnosis|HPE no|Note|$))', re.S),
        "Gross Examination": re.compile(r'Gross Examination\s*:\s*(.*?)(?=\s*(Microscopy|Diagnosis|HPE no|Note|$))', re.S),
        "Microscopy": re.compile(r'Microscopy\s*:\s*(.*?)(?=\s*(Diagnosis|HPE no|Note|$))', re.S),
        "Diagnosis": re.compile(r'Diagnosis\s*:\s*(.*?)(?=\s*(HPE no|Note|$))', re.S),
        "HPE no": re.compile(r'HPE no\s*:\s*(.*?)(?=\s*(Note|$))', re.S),
        "Note": re.compile(r'Note\s*[:-]\s*(.*?)(?=\s*$)', re.S),
    }

    for key, pattern in patterns.items():
        match = pattern.search(text)
        if match:
            value = match.group(1).strip()
            key_value_pairs[key] = value

    return key_value_pairs

def export_to_excel(key_value_pairs, image_name):
    """
    Export key-value pairs to an Excel file with the same name as the image file.
    """
    # Create a DataFrame from the key-value pairs
    df = pd.DataFrame(list(key_value_pairs.items()), columns=['Key', 'Value'])
    
    # Derive the Excel file name from the image file name
    excel_file_name = image_name.rsplit('.', 1)[0] + '.xlsx'
    
    # Create a new Workbook
    wb = Workbook()
    ws = wb.active
    
    # Append DataFrame to the worksheet
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)
    
    # Set alignment for all cells to wrap text and adjust row height
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)

    # Adjust column width based on the maximum length of content in each column
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    
    # Adjust row height to fit the content
    for row in ws.iter_rows():
        max_length = 0
        for cell in row:
            cell_length = len(str(cell.value).split('\n'))
            if cell_length > max_length:
                max_length = cell_length
        ws.row_dimensions[row[0].row].height = max_length * 15  # Adjust this multiplier as needed
    
    # Save the Workbook
    wb.save(excel_file_name)
    return excel_file_name

# Streamlit App
st.title("Text Extraction from Scanned Documents")

uploaded_file = st.file_uploader("Choose an image...", type=["jpg", "jpeg", "png"])

if uploaded_file is not None:
    # Display the uploaded image
    image = Image.open(uploaded_file)
    st.image(image, caption='Uploaded Image.', use_column_width=True)

    # Extract text from image
    text = extract_text_from_image(image)
    st.text_area("Extracted Text", text, height=200)

    # Extract key-value pairs
    key_value_pairs = extract_key_value_pairs(text)
    st.write("Extracted Key-Value Pairs:")
    st.json(key_value_pairs)

    # Export to Excel
    if st.button("Export to Excel"):
        excel_file_name = export_to_excel(key_value_pairs, uploaded_file.name)
        with open(excel_file_name, "rb") as file:
            btn = st.download_button(
                label="Download Excel",
                data=file,
                file_name=excel_file_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
